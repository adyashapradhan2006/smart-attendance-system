"""
csv_to_excel.py
===============
Run this to convert all attendance CSVs into a
nicely formatted Excel file with one sheet per date.

Usage:
    python csv_to_excel.py
"""

import pandas as pd
from pathlib import Path
from datetime import datetime

BASE_DIR       = Path(__file__).parent
ATTENDANCE_DIR = BASE_DIR / "attendance_logs"
OUTPUT_FILE    = BASE_DIR / "attendance_report.xlsx"

def convert():
    csv_files = sorted(ATTENDANCE_DIR.glob("attendance_*.csv"))
    if not csv_files:
        print("No attendance files found in attendance_logs/")
        return

    print(f"Found {len(csv_files)} attendance file(s)...")

    with pd.ExcelWriter(str(OUTPUT_FILE), engine="openpyxl") as writer:
        for csv_path in csv_files:
            date_str   = csv_path.stem.replace("attendance_", "")
            df         = pd.read_csv(str(csv_path))

            # Remove duplicate names — keep only FIRST entry per person per day
            df = df.drop_duplicates(subset=["Name"], keep="first")

            # Add serial number column
            df.insert(0, "S.No", range(1, len(df) + 1))

            # Write to Excel sheet named by date
            sheet_name = date_str  # e.g. "2026-03-03"
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Auto-size columns
            ws = writer.sheets[sheet_name]
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col) + 4
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 40)

            # Style the header row
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill("solid", fgColor="1a1a2e")
            for cell in ws[1]:
                cell.font      = Font(bold=True, color="00E5A0", size=11)
                cell.fill      = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Style data rows
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")

            print(f"  Sheet '{sheet_name}': {len(df)} unique person(s)")

    print(f"\nExcel report saved: {OUTPUT_FILE}")
    print("Open 'attendance_report.xlsx' in your smart_attendance folder!")

if __name__ == "__main__":
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl...")
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])

    convert()